#!/usr/bin/env python3
import argparse

# import os
import sys
from datetime import date
from datetime import datetime
from datetime import time
from datetime import timedelta
from pathlib import Path
from typing import cast
from typing import Optional
from typing import Sequence
from typing import TypeVar
import warnings

from lxml.etree import SubElement
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import cell as CELL
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

import package.utilities as utils
from package.tables import CH_Diary_Table
from package.table_sections import CH_DiaryDay_TableSection
from package.tables import CH_Diary_Table
from package.tables import create_contents_xml_element
from package.utilities import Excel, counters
from package.table_sections import SectionParent
from package.tables import Analysis_Table
from package.tables import WH_Diary_Table
from package.table_sections import WH_DiaryDay_TableSection
from package.utilities import format_date
from package.utilities import ID_Cell
from package.utilities import make_id_cells
from package.utilities import str_strip
from package.utilities import debug


warnings.filterwarnings(
    "ignore",
    "Data Validation.*",
)


Table_Type = TypeVar("Table_Type", bound=Analysis_Table)

Sections = list[tuple[str, str, str, Optional[SectionParent]]]


# override default openpyxl timedelta (duration) format
CELL.TIME_FORMATS[timedelta] = "[h].mm"

# In the westminster hall diary part, the chamber day number appearers in square brackets
# if the chamber diary part is created first we can store a set of key value pairs
# (dates and [chamber day numbers]) here and use it for the westminster hall diaryself.
# e.g. {2021-07-29: 1}
DATE_NUM_LOOK_UP: dict[date, int] = {}


#  We expect the following column headings in the Excel document
DAY = "Day"
DATE = "Date"
TIME = "Time"
SUBJECT1 = "Subject 1"
SUBJECT2 = "Subject 2"
TAGS = "Tags"
# DURATION = 'DurationFx'
DURATION = "Duration"
AAT = "AAT"

# these are the expected headings for the chamber sheet
# the order does not matter
CHAMBER_COLS = [DAY, DATE, TIME, SUBJECT1, SUBJECT2, TAGS, DURATION, AAT]
# and for westminster hall sheet
WH_COLS = [DAY, DATE, TIME, SUBJECT1, SUBJECT2, TAGS, DURATION]

CH_SHEET_TITLE = "Chamber"
WH_SHEET_TITLE = "Westminster Hall"


class WHRow:
    """Converts an openpyxl Excel row into a basic object.
    Must have title_index setup before first use.
    Fixes common but minor problems with the sessional diary Excel file.
    For example the DURATION field is expected to contain a datetime.timedelta
    but sometime contains either a datetime.datetime or a datetime.time.
    These are automatically converted."""

    title_index: dict[str, int] = {}

    def __init__(self, excel_row: Sequence[Cell]):

        # sequence and not Iterable because we use indexes (__getitem__)
        t_index = WHRow.title_index

        self.inner_init(excel_row, t_index)

    def inner_init(self, excel_row: Sequence[Cell], t_index: dict[str, int]):

        if not t_index:
            # maybe raise here instead
            print("Error: title_index not set up")
            exit()

        # expect an integer in the DAY column (this just counts sittings)
        self.day: int
        _day = excel_row[t_index[DAY]].value
        if isinstance(_day, int):
            self.day = _day
        else:
            print(excel_row[t_index[DAY]].coordinate, " has value ", _day)
            raise ValueError

        # expect a date object in the DATE column
        _date = excel_row[t_index[DATE]].value

        self.date: date
        if isinstance(_date, date):
            self.date = _date
        else:
            print(excel_row[t_index[DATE]].coordinate, " has value ", _date)
            raise ValueError

        # expect a time object in the TIME column
        # (but sometimes it is datetime so convert)
        self.time: time
        time_cell = excel_row[t_index[TIME]]
        _time = time_cell.value
        if isinstance(_time, time):
            self.time = _time
        else:
            if excel_row[t_index[TIME]].value is not None:
                print(excel_row[t_index[TIME]].coordinate, " has value ", _time)
            raise ValueError

        if isinstance(self.time, datetime):
            time_obj = self.time.time()
            print(f"There is a datetime at cell {time_cell.coordinate}: {self.time}")
            print(f"This has been converted to the following time: {time_obj}")
            self.time = time_obj

        self.subject1: str = str_strip(excel_row[t_index[SUBJECT1]].value)

        self.subject2: str = str_strip(excel_row[t_index[SUBJECT2]].value)

        self.tags = str_strip(excel_row[t_index[TAGS]].value)

        # expect timedelata in DURATION col
        # can also be datetime or time so try to convert
        duration_cell = excel_row[t_index[DURATION]]
        self.duration: timedelta
        if isinstance(duration_cell.value, datetime):
            # don't trust the datetime only the time
            time_obj = duration_cell.value.time()
            print(
                f"There is a datetime at cell {duration_cell.coordinate}:",
                str(duration_cell.value),
            )
            print(f"This has been converted to the following time: {time_obj}")
            self.duration = datetime.combine(date.min, time_obj) - datetime.min
        elif isinstance(duration_cell.value, time):
            self.duration = (
                datetime.combine(date.min, duration_cell.value) - datetime.min
            )
        elif not isinstance(duration_cell.value, timedelta):
            # TODO: log this
            # print(f'Problem in cell {duration_cell.coordinate}')
            self.duration = timedelta()

    def __str__(self):
        return (
            f"{self.day=} {self.date=} {self.time} {self.subject1=} "
            f"{self.subject2=} {self.tags=} {self.duration=}"
        )


class CHRow(WHRow):
    title_index: dict[str, int] = {}

    def __init__(self, excel_row: Sequence[Cell]):

        t_index = CHRow.title_index
        super().inner_init(excel_row, t_index)

        aat_cell = excel_row[t_index[AAT]]
        self.aat: timedelta
        if isinstance(aat_cell.value, datetime):
            # don't trust the datetime only the time
            time_obj = aat_cell.value.time()
            print(
                f"There is a datetime at cell {aat_cell.coordinate}:",
                str(aat_cell.value),
            )
            print(f"This has been converted to the following time: {time_obj}")
            # get timedelta
            self.aat = datetime.combine(date.min, time_obj) - datetime.min
        elif isinstance(aat_cell.value, time):
            # get timedelta
            self.aat = datetime.combine(date.min, aat_cell.value) - datetime.min
        elif not isinstance(aat_cell.value, timedelta):
            self.aat = timedelta()


class Sessional_Diary:
    def __init__(self, input_excel_file_path: str, no_excel: bool):

        self.input_workbook: Workbook = load_workbook(
            filename=input_excel_file_path, data_only=True, read_only=True
        )

        # if we require an output excel file
        if no_excel is False:
            Excel.out_wb = Workbook()  # new Excel workbook obj

    def check_chamber(self):
        try:
            cmbr_data = cast(Worksheet, self.input_workbook[CH_SHEET_TITLE])
        except Exception:
            print(
                'There is no "Chamber" worksheet in the Excel file.',
                "This sheet is required.",
            )
            exit()

        top_row = cmbr_data[1]
        CHRow.title_index = {item.value: i for i, item in enumerate(top_row)}

        if not set(CHAMBER_COLS).issubset(set(CHRow.title_index.keys())):
            expected_row_headings = '", "'.join(CHAMBER_COLS)
            print(
                f"Expected the following column titles "
                f"to be in the top row of the {CHAMBER_COLS} sheet\n",
                f'"{expected_row_headings}"',
            )

    def check_wh(self):
        try:
            wh_data = cast(Worksheet, self.input_workbook[WH_SHEET_TITLE])
        except Exception:
            print(
                f'There is no "{WH_SHEET_TITLE}" worksheet in the Excel file.',
                "This sheet is required.",
            )
            exit()

        top_row = wh_data[1]
        WHRow.title_index = {item.value: i for i, item in enumerate(top_row)}

        if not set(WH_COLS).issubset(set(WHRow.title_index.keys())):
            expected_row_headings = '", "'.join(WH_COLS)
            print(
                f"Expected the following column titles "
                f"to be in the top row of the {WH_SHEET_TITLE} sheet",
                f'"{expected_row_headings}"',
                f"Got {WHRow.title_index.keys()}",
                sep="\n",
            )

    def house_diary(self, output_folder_Path: Path):
        """Create an (indesign formatted) XML file for the house diary
        section of the Sessional diary."""

        output_file_Path = output_folder_Path.joinpath("House_Diary.xml")

        self.check_chamber()

        cmbr_data = cast(Worksheet, self.input_workbook[CH_SHEET_TITLE])

        table_ele = CH_Diary_Table(
            [
                ("Time", 35),
                ("Subject", 355),
                # ('Exit', 45),
                ("Duration", 45),
                ("After appointed time", 45),
            ]
        )

        previous_day = 1

        section = CH_DiaryDay_TableSection("This section should not be outputted")

        for c, excel_row in enumerate(cmbr_data.iter_rows(), start=1):
            if c == 1:
                continue

            if all(not cell.value for cell in excel_row[:10]):
                # skip over any blank rows
                continue

            try:
                entry = CHRow(excel_row)
            except (ValueError, AttributeError):
                print(f"Skipping row {c}")
                continue

            if c == 2:
                # create the first day section (to be added later to the diary table)
                section_title = (
                    f'{entry.day}.\u2002{entry.date.strftime("%A %d %B %Y")}'
                )
                section = table_ele.start_new_section(section_title)

            if entry.day != previous_day:
                previous_day = entry.day

                new_section_title = (
                    f'{entry.day}.\u2002{entry.date.strftime("%A %d %B %Y")}'
                )
                section = table_ele.start_new_section(new_section_title)

                # add the date and number to the lookup.
                # this is so this info can also be put in the WH table
                DATE_NUM_LOOK_UP[entry.date] = entry.day

            # need to add up all the durations
            table_ele.session_total_time += entry.duration
            # table_ele.session_total_after_moi += entry.aat
            table_ele.session_aat_total += entry.aat

            # there will be 4 cells per row
            cell = ID_Cell()

            # create a Bold element. Optionally can have non bold tail text
            bold = SubElement(cell, "Bold")
            bold.text = entry.subject1
            if entry.subject2:
                bold.tail = f": {entry.subject2}"  # this text will not be bold

            duration = entry.duration
            if duration == timedelta(seconds=0):
                duration = ""
            aat = entry.aat
            if aat == timedelta(seconds=0):
                aat = ""

            section.add_row(
                [entry.time.strftime("%H.%M"), cell, duration, aat],
                duration=entry.duration,
                aat=entry.aat,
            )

        # need to add the last table section
        table_ele.add_section(section)
        # if counters.section < 3:
        #     print(f"section={str(section)}")
        #     counters.section += 1

        # now output XML (for InDesign) file
        utils.write_xml(table_ele.xml_element, output_file_Path)

    def house_analysis(self, output_folder_Path: Path):

        output_file_Path = output_folder_Path.joinpath("House_An_Contents.xml")

        # we're only interested in the main data here
        main_data = cast(Worksheet, self.input_workbook[CH_SHEET_TITLE])

        # add heading elements to table
        table_ele = Analysis_Table(
            [("Date", 95), ("", 295), ("Duration", 45), ("After appointed time", 45)],
        )

        # parents
        # some tables have a parent e.g. 2 is the parent of 2a and 2b
        # parents are only referenced in the table of contents
        p_2 = SectionParent("2:\tGovernment bills")
        p_3 = SectionParent("3:\tPrivate Members’ bills")
        p_5 = SectionParent("5:\tGovernment motions")
        p_6 = SectionParent("6:\tOpposition business")
        p_8 = SectionParent("8:\tPrivate Members’ business (other than bills)")
        p_14 = SectionParent("14:\tBusiness when no Question before House")

        table_ele.start_new_section(
            "addresses",
            "1:\tAddresses other than Prayers",
            "1 Addresses other than Prayers",
            None,
        )

        # define the sub subsections
        # list of 4-tuples
        #   Item 1: key (used for access)
        #   Item 2: Title of the section for Indesign
        #   Item 3: Title of the section in Excel
        #   Item 4: parent (or none if this section will be top level)
        sections: Sections = [
            # the order matters!
            (
                "addresses",  # key
                "1:\tAddresses other than Prayers",  # title in Indesign
                "1 Addresses other than Prayers",  # title in Excel
                None,  # parent (None if top level)
            ),
            (
                "second_readings",
                "2a:\tGovernment Bills: Read a second time and committed to Public Bill Committee",
                "2a Govt Bills 2R & committed",
                p_2,
            ),
            (
                "cwh_bills",
                "2b:\tGovernment Bills: Read a second time and committed to "
                "Committee of the whole House (in whole or part)",
                "2b Govt Bill 2R & sent to CWH",
                p_2,
            ),
            (
                "cwh_2_bills",
                "2d:\tGovernment Bills: Committee of the whole House",
                "2d Govt Bills CWH",
                p_2,
            ),
            (
                "gov_bil_cons",
                "2e:\tGovernment Bills: Consideration",
                "2e Govt Bills Consideration",
                p_2,
            ),
            (
                "gov_bill_3rd",
                "2f:\tGovernment Bills: Third Reading",
                "2f Govt Bills 3R",
                p_2,
            ),
            (
                "gov_bill_lord_amend",
                "2g:\tGovernment Bills: Lord Amendments",
                "2g Lords Amendments",
                p_2,
            ),
            (
                "alloc_time",
                "2h:\tAllocation of time motions",
                "2h Allocation of time motions",
                p_2,
            ),
            (
                "gov_bill_other",
                "2i:\tGovernment Bills: Other Stages",
                "2i Govt Bills Other Stages",
                p_2,
            ),
            (
                "pmbs_2r",
                "3a:\tPrivate Members' Bills: Second Reading",
                "3a PMB 2R",
                p_3,
            ),
            (
                "pmbs_other",
                "3b:\tPrivate Members' Bills: Other Stages",
                "3b PMB Other stages",
                p_3,
            ),
            ("private_business", "4:\tPrivate Business", "4 Private Business", None),
            (
                "eu_docs",
                "5a:\tEuropean Union documents",
                "5a European Union documents",
                p_5,
            ),
            (
                "gov_motions",
                "5b:\tGovernment motions",
                "5b Government motions",
                p_5,
            ),
            (
                "gov_motions_gen",
                "5c:\tGovernment motions (General)",
                "5c Govt motions (General)",
                p_5,
            ),
            (
                "gen_debates",
                "5d:\tGovernment motions (General Debates)",
                "5d Govt motions (Gen Debates)",
                p_5,
            ),
            ("opposition_days", "6a:\tOpposition Days", "6a Opposition Days", p_6),
            (
                "oppo_motions_in_gov_time",
                "6b:\tOpposition motions in Government time",
                "6b Opp Motion in Govt time",
                p_6,
            ),
            (
                "backbench_business",
                "7: \tBackbench Business",
                "7 Backbench Business",
                None,
            ),
            (
                "pm_motion",
                "8a:\tPrivate Members' Motions",
                "8a Private Members' Motions",
                p_8,
            ),
            (
                "ten_min_motion",
                "8b:\tTen Minute Rule Motions",
                "8b Ten minute rules",
                p_8,
            ),
            (
                "emergency_debates",
                "8c:\tEmergency debates",
                "8c Emergency debates",
                p_8,
            ),
            (
                "adjournment_debates",
                "8d:\tAdjournment debates",
                "8d Adjournment debates",
                p_8,
            ),
            ("estimates", "9:\tEstimates", "9 Estimates", None),
            ("money", "10:\tMoney Resolutions", "10 Money Resolutions", None),
            ("ways_and_means", "11:\tWays and Means", "11 Ways and Means", None),
            (
                "affirmative_sis",
                "12:\tAffirmative Statutory Instruments",
                "12 Affirmative SIs",
                None,
            ),
            (
                "negative_sis",
                "13:\tNegative Statutory Instruments",
                "13 Negative SIs",
                None,
            ),
            ("questions", "14a:\tQuestions", "14a Questions", p_14),
            (
                "topical_questions",
                "14b:\tTopical Questions",
                "14b Topical Questions",
                p_14,
            ),
            (
                "urgent_questions",
                "14c:\tUrgent Questions",
                "14c Urgent Questions",
                p_14,
            ),
            ("statements", "14d:\tStatements", "14d Statements", p_14),
            (
                "business_statements",
                "14e:\tBusiness Statements",
                "14e Business Statements",
                p_14,
            ),
            (
                "committee_statements",
                "14f:\tCommittee Statements",
                "14f Committee Statements",
                p_14,
            ),
            (
                "app_for_emerg_debate",
                "14g:\tS.O. No. 24 Applications",
                "14g SO No 24 Applications",
                p_14,
            ),
            (
                "points_of_order",
                "14h:\tPoints of Order",
                "14h Points of Order",
                p_14,
            ),
            (
                "public_petitions",
                "14i:\tPublic Petitions",
                "14i Public Petitions",
                p_14,
            ),
            ("miscellaneous", "14j:\tMiscellaneous", "14j Miscellaneous", p_14),
            ("prayers", "15:\tDaily Prayers", "15 Daily Prayers", None),
        ]

        table_ele.add_new_sections(sections)

        # ------------------- loop over excel data ------------------- #

        for c, excel_row in enumerate(main_data.iter_rows(), start=1):

            if c == 1:
                # top row just has headings in
                continue

            if all(not v.value for v in excel_row[:9]):
                # skip any blank rows
                continue

            try:
                entry = CHRow(excel_row)
            except (ValueError, AttributeError):
                print(f"Skipping row {c}")
                continue

            forematted_date = format_date(entry.date)

            # col_time = Time.strip().lower()
            subject_lower = entry.subject1.lower()
            col_exit = entry.tags.lower()

            cells_vals = [forematted_date, entry.subject2, entry.duration, entry.aat]

            fullrow = (cells_vals, entry.duration, entry.aat)

            # Table 1 Addresses other than Prayers
            if subject_lower == "address":
                table_ele["addresses"].add_row(*fullrow)
            # Table 2a Government bills second reading
            if "[pmb]" not in col_exit:
                # here we have items that are not explicitly private members' bills
                if subject_lower == "second reading" and "pbc" in col_exit:
                    # gov bill second reading
                    table_ele["second_readings"].add_row(*fullrow)

                if "committee of the whole house" in subject_lower:
                    table_ele["cwh_2_bills"].add_row(*fullrow)
                if "consideration" in subject_lower:
                    # gov bill consideration
                    table_ele["gov_bil_cons"].add_row(*fullrow)
                if subject_lower == "third reading":
                    # gov bill third reading
                    table_ele["gov_bill_3rd"].add_row(*fullrow)
                if subject_lower == "lords amendments":
                    # gov bill lords amendments
                    table_ele["gov_bill_lord_amend"].add_row(*fullrow)
                gov_bill_other_subs = (
                    "second and third reading",
                    "money resolution",
                    "lords amendments",
                )
                if (
                    "legislative grand committee" in subject_lower
                    or subject_lower in gov_bill_other_subs
                ):
                    table_ele["gov_bill_other"].add_row(*fullrow)

            if (
                subject_lower == "second reading"
                and "committee of the whole house" in entry.subject2.lower()
            ):
                table_ele["cwh_bills"].add_row(*fullrow)

            if subject_lower.lower() in ("general debate", "general motion"):
                table_ele["alloc_time"].add_row(*fullrow)

            if "[pmb]" in col_exit:
                if subject_lower == "second reading":
                    # private members' bills second reading
                    table_ele["pmbs_2r"].add_row(*fullrow)
                elif subject_lower not in (
                    "ten minute rule motion",
                    "point of order",
                    "remaining orders",
                ):
                    # private members' bills other
                    # this does not include ten minute rules
                    table_ele["pmbs_other"].add_row(*fullrow)

            if "private business" in subject_lower:
                table_ele["private_business"].add_row(*fullrow)

            if subject_lower == "eu documents":
                table_ele["eu_docs"].add_row(*fullrow)

            if subject_lower in (
                "government motion",
                "government motions",
                "business motion",
            ):
                table_ele["gov_motions"].add_row(*fullrow)

            if subject_lower in ("general motion"):
                table_ele["gov_motions_gen"].add_row(*fullrow)

            if subject_lower == "general debate":
                table_ele["gen_debates"].add_row(*fullrow)

            if subject_lower == "opposition day":
                table_ele["opposition_days"].add_row(*fullrow)

            if subject_lower == "opposition motion in government time":
                table_ele["oppo_motions_in_gov_time"].add_row(*fullrow)
            if subject_lower == "backbench business":
                table_ele["backbench_business"].add_row(*fullrow)
            if subject_lower in (
                "private member's motion",
                "private member’s motion",
                "private members' motion",
            ):
                table_ele["pm_motion"].add_row(*fullrow)
            if subject_lower == "ten minute rule motion":
                table_ele["ten_min_motion"].add_row(*fullrow)
            if "no. 24 debate" in subject_lower:
                table_ele["emergency_debates"].add_row(*fullrow)
            if "adjournment" in subject_lower:
                table_ele["adjournment_debates"].add_row(*fullrow)
            if subject_lower == "estimates day":
                table_ele["estimates"].add_row(*fullrow)
            if subject_lower == "money resolution":
                table_ele["money"].add_row(*fullrow)
            if subject_lower == "ways and means":
                table_ele["ways_and_means"].add_row(*fullrow)
            if "affirmative" in subject_lower:
                table_ele["affirmative_sis"].add_row(*fullrow)
            if subject_lower == "negative statutory instrument":
                table_ele["negative_sis"].add_row(*fullrow)
            if subject_lower == "questions":
                table_ele["questions"].add_row(*fullrow)
            if subject_lower == "topical questions":
                table_ele["topical_questions"].add_row(*fullrow)
            if subject_lower in ("urgent question", "urgent questions"):
                table_ele["urgent_questions"].add_row(*fullrow)
            if subject_lower == "statement":
                table_ele["statements"].add_row(*fullrow)
            if subject_lower == "business statement":
                table_ele["business_statements"].add_row(*fullrow)
            if "committee statement" in subject_lower:
                table_ele["committee_statements"].add_row(*fullrow)
            if "no. 24 application" in subject_lower:
                table_ele["app_for_emerg_debate"].add_row(*fullrow)
            if subject_lower in ("point of order", "points of order"):
                table_ele["points_of_order"].add_row(*fullrow)
            if "public petition" in subject_lower:
                table_ele["public_petitions"].add_row(*fullrow)
            if subject_lower == "prayers":
                # prayers are not itemised
                # table_ele['prayers'].add_row(*fullrow)
                table_ele["prayers"].duration += entry.duration
                table_ele["prayers"].after_appointed_time += entry.aat

            miscellaneous_options = (
                "tributes",
                "election of a speaker",
                "suspension",
                "observation of a minute's silence",
                "personal statement",
                "presentation of private members' bills",
            )
            if (
                subject_lower in miscellaneous_options
                or "message to attend the lords" in subject_lower
            ):

                # for Miscellaneous we will also include stuff in col_subject3
                misc_cells = [
                    forematted_date,
                    ": ".join([entry.subject1, entry.subject2]).rstrip(": "),
                    entry.duration,
                    entry.aat,
                ]
                table_ele["miscellaneous"].add_row(
                    misc_cells, entry.duration, entry.aat
                )

        previous_table_sec_parent: Optional[SectionParent] = None

        for table_section in table_ele.sections.values():
            if table_section.parent != previous_table_sec_parent:
                # if there is a section with a new parent we will put a
                # new subhead row into the table This will probably
                # make logical sense in the table and will definitely
                # make it easier to genareate the table of contents
                # in InDesign
                previous_table_sec_parent = table_section.parent
                if table_section.parent is not None:
                    # add a sub head row
                    table_ele.add_table_sub_head(table_section.parent.title)

            if "daily prayers" in table_section.title.lower():
                # Daily prayers is left blank on purpose and still needs to be added
                table_ele.add_section(table_section)
            elif len(table_section) > 0:
                table_ele.add_section(table_section)
            else:
                # add empty table sections but put nil in.
                cells_vals = ["Nil", "", "", ""]
                table_section.add_row(cells_vals, timedelta(), timedelta())
                table_ele.add_section(table_section)

        # now create XML for InDesign
        # create root element
        utils.write_xml(
            table_ele.xml_element, output_folder_Path.joinpath("House_Analysis.xml")
        )

        contents_xml = create_contents_xml_element(table_ele)
        utils.write_xml(contents_xml, output_file_Path)

    def wh_diary(self, output_folder_Path: Path):

        output_file_Path = output_folder_Path.joinpath("WH_diary.xml")

        self.check_wh()

        table_ele = WH_Diary_Table(
            [("Time", 35), ("Subject", 400), ("Duration", 45)],
        )

        if len(DATE_NUM_LOOK_UP) == 0:
            print(
                "Data for the chamber has not yet been processed so the chamber number will"
                " not be put in the westminstar hall table. The square brackets will"
                " instead be left blank."
            )

        wh_data = cast(Worksheet, self.input_workbook[WH_SHEET_TITLE])

        previous_day = 1

        section = WH_DiaryDay_TableSection("This section should not be outputted")

        for c, excel_row in enumerate(wh_data.iter_rows(), start=1):

            if c == 1:
                continue

            try:
                entry = WHRow(excel_row)
            except (ValueError, AttributeError):
                continue

            if c == 2:
                chamber_daynum = DATE_NUM_LOOK_UP.get(entry.date, "")
                sec_title = (
                    f"{entry.day}.\u2002[{chamber_daynum}]"
                    f'\u2002{entry.date.strftime("%A %d %B %Y")}'
                )
                section = table_ele.start_new_section(sec_title)
                # table_sections.append(WH_DiaryDay_TableSection(sec_title))

            if entry.day != previous_day:
                previous_day = entry.day

                # table_ele.add_section(table_sections[-1], session_total_time)
                # table_sections[-1].add_to(table_ele, session_total_time)

                # if the chamber diary has already been created the global
                # dictionary, `DATE_NUM_LOOK_UP` will have been populated  with
                # datetime.date objs as the keys and Integers as values. If the
                # chamber diary has not already been created or if westminster
                # hall sat on a day where the chamber did not sit, we may have
                # empty square brackets.
                chamber_daynum = DATE_NUM_LOOK_UP.get(entry.date, "")

                sec_title = (
                    f"{entry.day}.\u2002[{chamber_daynum}]"
                    f'\u2002{entry.date.strftime("%A %d %B %Y")}'
                )

                section = table_ele.start_new_section(sec_title)
                # table_sections.append(WH_DiaryDay_TableSection(sec_title))

            # need to add up all the durations
            table_ele.session_total_time += entry.duration

            # there will be 3 cells per row
            cell = ID_Cell()
            if entry.subject1:
                bold = SubElement(cell, "Bold")
                bold.text = entry.subject1
                if entry.subject2:
                    bold.tail = f": {entry.subject2}"
                cells = make_id_cells(
                    [entry.time.strftime("%H.%M"), cell, entry.duration]
                )
                section.add_row(cells, entry.duration)

        # last table section will not have been added in the above loop
        table_ele.add_section(section)
        # table_sections[-1].add_to(table_ele, session_total_time)

        # now output the XML for InDesign

        utils.write_xml(table_ele.xml_element, output_file_Path)

    def wh_analysis(self, output_folder_Path: Path):

        output_file_Path = output_folder_Path.joinpath("WH_Analysis.xml")

        wh_data = cast(Worksheet, self.input_workbook["Westminster Hall"])

        table_ele = Analysis_Table(  # new table element with headings
            [("Date", 95), ("Detail", 340), ("Duration", 45)],
        )

        # parents
        p_1 = SectionParent("1:\tPrivate Members")

        sections: Sections = [
            # the order matters!
            ("private", "1a:\tPrivate Members’ Debates", "WH1 Members debates", p_1),
            (
                "bbcom",
                "1b:\tPrivate Members’ (Backbench Business Committee recommended) Debates",
                "WH2 BBCom debates",
                p_1,
            ),
            (
                "liaison",
                "2:\tLiaison Committee Debates",
                "WH3 Liaison Com debates",
                None,
            ),
            ("e_petition", "3:\tDebates on e-Petitions", "WH4 e-Petitions", None),
            ("suspension", "4:\tSuspensions", "WH5 Suspensions", None),
            ("miscellaneous", "5:\tMiscellaneous", "WH6 Miscellaneous", None),
            ("statements", "6:\tStatements", "WH7 Statements", None),
        ]

        table_ele.add_new_sections(sections)

        for c, excel_row in enumerate(wh_data.iter_rows(), start=1):
            if c == 1:  # note start=1
                # skip first row
                continue

            if all(bool(v) is False for v in excel_row[:8]):
                # skip any blank rows
                continue

            try:
                entry = WHRow(excel_row)
            except (ValueError, AttributeError):
                print(f"WH Skipping row: {excel_row}")
                continue

            forematted_date = format_date(entry.date)

            cells_vals = [
                forematted_date,
                entry.subject2,
                entry.duration,
            ]
            fullrow = (cells_vals, entry.duration)

            # logic to determin in which sections entries go
            if entry.subject1 in (
                "Debate (Private Member’s)",
                "Debate (Private Member's)",
            ):
                table_ele["private"].add_row(*fullrow)

            elif entry.subject1 in (
                "Debate (BBCom recommended)",
                "Debate (BBCom)",
                "Debate (BBBCom)",
            ):
                table_ele["bbcom"].add_row(*fullrow)

            elif entry.subject1 in ("Debate (Liaison Committee)",):
                table_ele["liaison"].add_row(*fullrow)

            elif entry.subject1 in ("Petition", "Petitions"):
                table_ele["e_petition"].add_row(*fullrow)

            elif entry.subject1 == "Suspension" and entry.tags not in (
                "[Questions]",
                "[Question]",
            ):
                table_ele["suspension"].add_row(*fullrow)

            elif entry.subject1 in ("Committee Statement",):
                table_ele["statements"].add_row(*fullrow)

            elif entry.subject1 in (
                "Time limit",
                "Time Limit",
                "Observation of a period of silence",
            ):
                table_ele["miscellaneous"].add_row(*fullrow)
            else:
                # print(f"WH entry in column {c} not added to any Analysis section")
                # print(str(entry))
                pass

        previous_table_sec_parent = None
        for table_section in table_ele.sections.values():
            if table_section.parent != previous_table_sec_parent:
                # if there is a section with a new parent we will put a
                # new subhead row into the table This will probably
                # make logical sense in the table and will definitely
                # make it easier to genareate the table of contents
                # in InDesign
                previous_table_sec_parent = table_section.parent
                if table_section.parent is not None:
                    # add a sub head row
                    table_ele.add_table_sub_head(table_section.parent.title)
            if len(table_section) > 0:
                table_ele.add_section(table_section)
            else:
                # adding empty table sections but put nil in.
                cells_vals = ["Nil", "", ""]
                table_section.add_row(cells_vals, timedelta())
                table_ele.add_section(table_section)

        # create XML for indesign
        utils.write_xml(table_ele.xml_element, output_file_Path)

        # create XML for contents for InDesign
        contents_xml = create_contents_xml_element(table_ele)
        utils.write_xml(contents_xml, output_folder_Path.joinpath("WH_An_Contents.xml"))


def main():

    if len(sys.argv) > 1:
        # do cmd line version
        parser = argparse.ArgumentParser(
            description="Process Sessional diary Excel and create XML for InDesign"
        )

        parser.add_argument(
            "input",
            metavar="input_file",
            type=open,
            help="File path to the Excel file you wish to process. "
            "If there are spaces in the path you must use quotes.",
        )

        parser.add_argument(
            "--no-excel",
            action="store_true",
            help="Use this flag if you want do not want to output an excel file.",
        )

        parser.add_argument(
            "--include-only",
            type=str,
            choices=["chamber", "wh"],
            help="Use this option if you want to include *only* "
            "one section (e.g. just the Chamber section) "
            "rather than both sections",
        )

        parser.add_argument(
            "--debug",
            action="store_true",
            help="Export indented xml to aid debugging ",
        )

        args = parser.parse_args(sys.argv[1:])

        if args.debug:
            debug.debug = True

        if args.include_only == "chamber":
            run(args.input.name, include_wh=False, no_excel=args.no_excel)
        elif args.include_only == "wh":
            run(args.input.name, include_chamber=False, no_excel=args.no_excel)
        else:
            run(args.input.name, no_excel=args.no_excel)

    else:
        # run the GUI version
        from package import gui

        gui.mainloop(run_callback=run)


def run(
    excel_file_path: str,
    output_folder_path: str = "",
    include_chamber: bool = True,
    include_wh: bool = True,
    no_excel: bool = False,
):

    if output_folder_path:
        output_folder_Path = Path(output_folder_path)
    else:
        output_folder_Path = Path(excel_file_path).parent

    sessional_diary = Sessional_Diary(excel_file_path, no_excel)

    if include_chamber:
        # create house diary
        sessional_diary.house_diary(output_folder_Path)

        # create house analysis
        sessional_diary.house_analysis(output_folder_Path)

    if include_wh:
        # crete Westminster hall diary
        sessional_diary.wh_diary(output_folder_Path)

        # create Westminster hall analysis
        sessional_diary.wh_analysis(output_folder_Path)

    # remove the default sheet
    if Excel.out_wb is not None:
        del Excel.out_wb["Sheet"]
        Excel.out_wb.save(filename=str(output_folder_Path.joinpath("Analysis.xlsx")))


# def id_table(
#     list_of_tuples: list[tuple[str, int]], table_class: Type[Table_Type]
# ) -> Table_Type:
#     """Takes a list of 2 tuples of table header and cell widths"""

#     # table element
#     # the table_class is based on etree.ElementBase
#     # name of the xml element will default to the name of the class
#     table_ele: Table_Type
#     table_ele = table_class(  # type: ignore
#         nsmap=NS_MAP,
#         attrib={
#             AID + "table": "table",
#             AID + "tcols": str(len(list_of_tuples)),
#             AID5 + "tablestyle": "Part1Table",
#         },
#     )
#     table_ele.increment_rows()

#     # add heading elements to table
#     for item in list_of_tuples:
#         heading = SubElement(
#             table_ele,
#             "Cell",
#             attrib={
#                 AID + "table": "cell",
#                 AID + "theader": "",
#                 AID + "ccolwidth": str(item[1]),
#             },
#         )
#         heading.text = item[0]

#     return table_ele


if __name__ == "__main__":
    main()
